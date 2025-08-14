import pandas as pd
from pathlib import Path
import re
from openpyxl import Workbook
from openpyxl.styles import Alignment

# 🔹 Función para limpiar la descripción quitando "Test X"
def limpiar_descripcion(texto):
    return re.sub(r"Test\s*\d+\s*-", "", texto).strip() if texto else texto

# Clase para representar un solo test
class LogTest:
    def __init__(self, archivo, test_num, descripcion, serie_pcb, limite_inf,
                 limite_sup, medida, unidades, temp_ini, temp_fin, resultado):
        self.archivo = archivo
        self.test_num = int(test_num) if test_num and test_num.isdigit() else None
        self.descripcion = descripcion
        self.serie_pcb = serie_pcb
        self.limite_inf = limite_inf
        self.limite_sup = limite_sup
        self.medida = medida
        self.unidades = unidades
        self.temp_ini = temp_ini
        self.temp_fin = temp_fin
        self.resultado = resultado

# Clase para analiza un archivo de log
class LogParser:
    def __init__(self, path_archivo):
        self.path_archivo = path_archivo

    def parse(self):
        tests = []
        with self.path_archivo.open("r", encoding="utf-8") as archivo:
            contenido = archivo.read()

            bloques = re.split(r"(?=Test Description:)", contenido)

            for bloque in bloques:
                if "Test Description:" not in bloque:
                    continue

                test_desc = re.search(r"Test Description:\s*(.*)", bloque)
                test_num = re.search(r"Test (\d+)", bloque)
                pcb_sn = re.search(r"PCB Serial Number:\s*(.*)", bloque)
                lower = re.search(r"Test Lower Limit:\s*(.*)", bloque)
                upper = re.search(r"Test Upper Limit:\s*(.*)", bloque)
                measure = re.search(r"Test Measurement:\s*(.*)", bloque)
                units = re.search(r"Units:\s*(.*)", bloque)
                temp_start = re.search(r"Starting Temperature.*:\s*(.*)", bloque)
                temp_end = re.search(r"Ending Temperature.*:\s*(.*)", bloque)
                result = re.search(r"Test Result:\s*(.*)", bloque)

                campos = [test_desc, test_num, pcb_sn, lower, upper, measure, units, temp_start, temp_end, result]
                if any(c and c.group(1).strip() == "N/A" for c in campos):
                    continue

                descripcion_limpia = limpiar_descripcion(test_desc.group(1).strip() if test_desc else None)

                test = LogTest(
                    archivo=self.path_archivo.name,
                    test_num=test_num.group(1) if test_num else None,
                    descripcion=descripcion_limpia,
                    serie_pcb=pcb_sn.group(1).strip() if pcb_sn else None,
                    limite_inf=lower.group(1).strip() if lower else None,
                    limite_sup=upper.group(1).strip() if upper else None,
                    medida=measure.group(1).strip() if measure else None,
                    unidades=units.group(1).strip() if units else None,
                    temp_ini=temp_start.group(1).strip() if temp_start else None,
                    temp_fin=temp_end.group(1).strip() if temp_end else None,
                    resultado=result.group(1).strip() if result else None,
                )

                tests.append(test)

        return tests

# Clase para procesar múltiples archivos y exportar resultados
class LogProcessor:
    def __init__(self, carpeta_logs):
        self.carpeta_logs = Path(carpeta_logs)
        self.tests = []

    def procesar_logs(self):
        if not self.carpeta_logs.exists() or not self.carpeta_logs.is_dir():
            print("❌ Carpeta no encontrada:", self.carpeta_logs)
            return

        for archivo_log in self.carpeta_logs.glob("*.txt"):
            parser = LogParser(archivo_log)
            self.tests.extend(parser.parse())

    def exportar_excel_multinivel(self, nombre_archivo="tests_excel_multinivel.xlsx"):
        agrupados = {}
        for test in self.tests:
            clave = (test.descripcion, test.limite_inf, test.limite_sup)
            agrupados.setdefault(clave, {}).setdefault(test.serie_pcb, []).append(test.medida)

        wb = Workbook()
        ws = wb.active
        ws.title = "Resultados"

        encabezado_fijo = ["Test", "LowLimit", "HighLimit"]
        todos_pcbs = sorted({pcb for tests_por_pcb in agrupados.values() for pcb in tests_por_pcb.keys()})

        header_row_1 = encabezado_fijo[:] + [""]
        header_row_2 = [""] * len(encabezado_fijo) + [""]
        header_row_3 = [""] * len(encabezado_fijo) + [""]

        for pcb in todos_pcbs:
            header_row_1.extend(["Serial Number"] * 3 + [""])
            header_row_2.extend([pcb] * 3 + [""])
            header_row_3.extend(["Trial1", "Trial2", "Trial3"] + [""])

        ws.append(header_row_1)
        ws.append(header_row_2)
        ws.append(header_row_3)

        col = len(encabezado_fijo) + 2
        for pcb in todos_pcbs:
            ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col+2)
            ws.merge_cells(start_row=2, start_column=col, end_row=2, end_column=col+2)
            for i in range(3):
                ws.cell(row=3, column=col+i).alignment = Alignment(horizontal="center")
            col += 4  # 3 trial columns + 1 empty column

        for clave, pcb_tests in agrupados.items():
            descripcion, limite_inf, limite_sup = clave
            fila = [descripcion, limite_inf, limite_sup, ""]
            for pcb in todos_pcbs:
                mediciones = pcb_tests.get(pcb, [])
                fila.extend(mediciones[:3] + [""] * (3 - len(mediciones)) + [""])
            ws.append(fila)

        wb.save(nombre_archivo)
        print(f"✅ Archivo Excel exportado como '{nombre_archivo}' con separaciones entre bloques de serial numbers.")

# Función principal
def main():
    ruta = r"C:\Users\3002975\Documents\GitHub\leer_log_prototipo-10\Py Proyect\leer_log_py"
    procesador = LogProcessor(ruta)
    procesador.procesar_logs()
    procesador.exportar_excel_multinivel()

if __name__ == "__main__":
    main()