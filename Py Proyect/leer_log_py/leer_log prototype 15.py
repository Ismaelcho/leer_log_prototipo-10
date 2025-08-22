import re
from pathlib import Path
from tkinter import Tk, filedialog, messagebox
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill

def limpiar_descripcion(texto):
    return re.sub(r"Test\s*\d+\s*-", "", texto).strip() if texto else texto

class LogTest:
    def __init__(self, archivo, test_num, descripcion, serie_pcb, limite_inf,
                 limite_sup, medida, unidades, temp_ini, temp_fin, resultado):
        self.archivo = archivo
        self.test_num = int(test_num) if test_num and test_num.isdigit() else None
        self.descripcion = descripcion
        self.serie_pcb = serie_pcb
        self.limite_inf = limite_inf
        self.limite_sup = limite_sup
        self.medida = medida
        self.unidades = unidades
        self.temp_ini = temp_ini
        self.temp_fin = temp_fin
        self.resultado = resultado

class LogParser:
    def __init__(self, path_archivo):
        self.path_archivo = path_archivo

    def parse(self):
        tests = []
        with self.path_archivo.open("r", encoding="utf-8") as archivo:
            contenido = archivo.read()
            bloques = re.split(r"(?=Test Description:)", contenido)
            for bloque in bloques:
                if "Test Description:" not in bloque:
                    continue

                test_desc = re.search(r"Test Description:\s*(.*)", bloque)
                test_num = re.search(r"Test (\d+)", bloque)
                pcb_sn = re.search(r"PCB Serial Number:\s*(.*)", bloque)
                lower = re.search(r"Test Lower Limit:\s*(.*)", bloque)
                upper = re.search(r"Test Upper Limit:\s*(.*)", bloque)
                measure = re.search(r"Test Measurement:\s*(.*)", bloque)
                units = re.search(r"Units:\s*(.*)", bloque)
                temp_start = re.search(r"Starting Temperature.*:\s*(.*)", bloque)
                temp_end = re.search(r"Ending Temperature.*:\s*(.*)", bloque)
                result = re.search(r"Test Result:\s*(.*)", bloque)

                campos = [test_desc, test_num, pcb_sn, lower, upper, measure, units, temp_start, temp_end, result]
                if any(c and c.group(1).strip() == "N/A" for c in campos):
                    continue

                descripcion_limpia = limpiar_descripcion(test_desc.group(1).strip() if test_desc else None)

                test = LogTest(
                    archivo=self.path_archivo.name,
                    test_num=test_num.group(1) if test_num else None,
                    descripcion=descripcion_limpia,
                    serie_pcb=pcb_sn.group(1).strip() if pcb_sn else None,
                    limite_inf=lower.group(1).strip() if lower else None,
                    limite_sup=upper.group(1).strip() if upper else None,
                    medida=measure.group(1).strip() if measure else None,
                    unidades=units.group(1).strip() if units else None,
                    temp_ini=temp_start.group(1).strip() if temp_start else None,
                    temp_fin=temp_end.group(1).strip() if temp_end else None,
                    resultado=result.group(1).strip() if result else None,
                )
                tests.append(test)
        return tests

class LogProcessor:
    def __init__(self, carpeta_pcb):
        self.carpeta_pcb = Path(carpeta_pcb)
        self.tests = []

    def procesar_logs(self):
        for archivo_log in self.carpeta_pcb.glob("*.txt"):
            parser = LogParser(archivo_log)
            self.tests.extend(parser.parse())

        for subcarpeta in self.carpeta_pcb.iterdir():
            if subcarpeta.is_dir():
                for archivo_log in subcarpeta.glob("*.txt"):
                    parser = LogParser(archivo_log)
                    self.tests.extend(parser.parse())

    def agregar_hoja_excel(self, wb, nombre_hoja):
        rojo = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        agrupados = {}
        for test in self.tests:
            clave = (test.descripcion, test.limite_inf, test.limite_sup)
            agrupados.setdefault(clave, {}).setdefault(test.serie_pcb, []).append(test.medida)

        ws = wb.create_sheet(title=nombre_hoja)

        encabezado_fijo = ["Test", "LowLimit", "HighLimit"]
        todos_pcbs = sorted({pcb for tests_por_pcb in agrupados.values() for pcb in tests_por_pcb.keys()})

        header_row_1 = encabezado_fijo[:] + [""]
        header_row_2 = [""] * len(encabezado_fijo) + [""]
        header_row_3 = [""] * len(encabezado_fijo) + [""]

        for pcb in todos_pcbs:
            header_row_1.extend(["Serial Number"] * 3 + [""])
            header_row_2.extend([pcb] * 3 + [""])
            header_row_3.extend(["Trial1", "Trial2", "Trial3"] + [""])

        ws.append(header_row_1)
        ws.append(header_row_2)
        ws.append(header_row_3)

        col = len(encabezado_fijo) + 2
        for pcb in todos_pcbs:
            ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col+2)
            ws.merge_cells(start_row=2, start_column=col, end_row=2, end_column=col+2)
            for i in range(3):
                ws.cell(row=3, column=col+i).alignment = Alignment(horizontal="center")
            col += 4

        for clave, pcb_tests in agrupados.items():
            descripcion, limite_inf, limite_sup = clave
            fila = [descripcion, limite_inf, limite_sup, ""]
            for pcb in todos_pcbs:
                mediciones = pcb_tests.get(pcb, [])
                fila.extend(mediciones[:3] + [""] * (3 - len(mediciones)) + [""])
            ws.append(fila)

            fila_idx = ws.max_row
            col_idx = len(encabezado_fijo) + 2
            for pcb in todos_pcbs:
                mediciones = pcb_tests.get(pcb, [])
                for i in range(3):
                    if i < len(mediciones):
                        try:
                            valor = float(mediciones[i])
                            limite_inf_f = float(limite_inf)
                            limite_sup_f = float(limite_sup)
                            if valor < limite_inf_f or valor > limite_sup_f:
                                ws.cell(row=fila_idx, column=col_idx + i).fill = rojo
                        except ValueError:
                            pass
                col_idx += 4

def main():
    root = Tk()
    root.withdraw()

    while True:
        carpeta_seleccionada = filedialog.askdirectory(title="Seleccione la carpeta con todas las PCB o especifique solo una")
        if not carpeta_seleccionada:
            salir = messagebox.askyesno("¿Desea salir?", "No se seleccionó ninguna carpeta.\n¿Desea cerrar el programa?")
            if salir:
                print("🚪 Programa finalizado por el usuario.")
                return
            else:
                continue

        carpeta_seleccionada = Path(carpeta_seleccionada)
        wb = Workbook()
        wb.remove(wb.active)

        subcarpetas = [f for f in carpeta_seleccionada.iterdir() if f.is_dir()]
        es_milog = any(
            any(sub.glob("*.txt")) or any((sub / sc).glob("*.txt"))
            for sub in subcarpetas for sc in sub.iterdir() if sc.is_dir()
        )

        total_logs = 0
        logs_invalidos = 0
        hojas_generadas = 0

        if es_milog:
            for carpeta_pcb in subcarpetas:
                if carpeta_pcb.is_dir():
                    procesador = LogProcessor(carpeta_pcb)
                    procesador.procesar_logs()
                    total_logs += len(procesador.tests)
                    if procesador.tests:
                        procesador.agregar_hoja_excel(wb, carpeta_pcb.name)
                        hojas_generadas += 1
                    else:
                        logs_invalidos += 1
        else:
            procesador = LogProcessor(carpeta_seleccionada)
            procesador.procesar_logs()
            total_logs += len(procesador.tests)
            if procesador.tests:
                procesador.agregar_hoja_excel(wb, carpeta_seleccionada.name)
                hojas_generadas += 1
            else:
                logs_invalidos += 1

        if hojas_generadas == 0:
            messagebox.showinfo("Sin datos", "No se encontraron datos válidos para exportar.")
            continuar = messagebox.askyesno("¿Desea intentar con otra carpeta?", "¿Desea seleccionar otra carpeta para procesar?")
            if continuar:
                continue
            else:
                print("🚪 Programa finalizado por el usuario.")
                return

        nombre_archivo = filedialog.asksaveasfilename(
            title="¿Dónde desea guardar su archivo Excel?",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile=f"{carpeta_seleccionada.name}_tests.xlsx"
        )

        if not nombre_archivo:
            cancelar_guardado = messagebox.askyesno("¿Desea intentar con otra carpeta?", "No se seleccionó ubicación para guardar el archivo.\n¿Desea seleccionar otra carpeta para procesar?")
            if cancelar_guardado:
                continue
            else:
                print("🚪 Programa finalizado por el usuario.")
                return

        while True:
            try:
                wb.save(nombre_archivo)
                print(f"✅ Archivo Excel exportado como '{nombre_archivo}'")
                break
            except PermissionError:
                reintentar = messagebox.askretrycancel(
                    "Archivo en uso",
                    f"No se pudo guardar el archivo '{nombre_archivo}' porque está abierto.\n\nPor favor, cierra el archivo Excel y elige 'Reintentar' para intentar nuevamente o 'Cancelar' para salir."
                )
                if not reintentar:
                    messagebox.showinfo("Cancelado", "La exportación fue cancelada. El archivo no se guardó.")
                    return

        messagebox.showinfo(
            "Resumen de ejecución",
            f"✅ Informacion Procesada: {total_logs}\n"
            f"⚠️ N/A o vacíos Detectados: {logs_invalidos}\n"
            f"📄 Hojas de Excel Generadas: {hojas_generadas}"
        )

        repetir = messagebox.askyesno("¿Procesar otra carpeta?", "¿Desea procesar otra carpeta PCB?")
        if not repetir:
            print("✅ Proceso finalizado correctamente.")
            break

if __name__ == "__main__":
    main()